"""
formatter.py

Workbook formatting and export layer.

Responsibilities:
    - Write all output sheets to a timestamped Excel workbook
    - Apply consistent header styling (bold, green fill, centred)
    - Freeze the header row on every sheet
    - Add auto-filter dropdowns
    - Auto-fit column widths (capped at 60 characters)
    - Apply number formatting to money columns
    - Highlight warning rows in the Warnings sheet
    - Write the Pivot2 title row ("Sum of Provision") before its headers

BUGS FIXED vs. original formatter.py:
    1. _number_format targeted "business_amount" — a column that is never
       written to the output sheets (it only exists internally in the
       enriched dataframe).  Fixed to target the actual output column
       names: "Provision" (Sheet1) and "Total" (Pivot1, Pivot2).

    2. Pivot2 requires a title row ("Sum of Provision") at row 3 and the
       column headers at row 4, with rows 1-2 blank, matching the expected
       Excel structure.  Added _write_pivot2_with_header() to handle this.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.config import Config
from core.logger import AppLogger


class FormatterError(Exception):
    """Raised when workbook export fails."""
    pass


class Formatter:
    """
    Excel workbook export and formatting engine.
    """

    # Column names that contain money values — these get number formatting
    MONEY_COLUMNS = {"Provision", "Total", "business_amount"}

    def __init__(self, config: Config, logger: AppLogger) -> None:
        self.config = config
        self.logger = logger

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export(
        self,
        output_dir: Path,
        sheets: Dict[str, pd.DataFrame],
        warnings_df: Optional[pd.DataFrame],
        metadata_df: Optional[pd.DataFrame],
    ) -> Path:
        """
        Write all sheets to a new Excel workbook and apply formatting.

        Args:
            output_dir:   Directory where the workbook is created.
            sheets:       Dict of sheet_name → DataFrame.
            warnings_df:  Rows with money warnings (or empty DataFrame).
            metadata_df:  Run statistics table (or None).

        Returns:
            Path to the generated workbook.
        """

        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            workbook_path = output_dir / self._generate_name()

            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                self._write_main_sheets(writer, sheets)
                self._write_warning_sheet(writer, warnings_df)
                self._write_metadata_sheet(writer, metadata_df)

            # Post-processing: openpyxl styling pass
            self._format_workbook(workbook_path, sheets)

            self.logger.success(f"Workbook exported: {workbook_path.name}")
            return workbook_path

        except Exception as exc:
            raise FormatterError(f"Workbook export failed: {exc}") from exc

    # ------------------------------------------------------------------
    # Sheet writing
    # ------------------------------------------------------------------

    def _generate_name(self) -> str:
        prefix = self.config.get("output", "output_prefix", "SAP_Automation")
        fmt    = self.config.get("output", "timestamp_format", "%Y%m%d_%H%M%S")
        return f"{prefix}_{datetime.now().strftime(fmt)}.xlsx"

    def _write_main_sheets(self, writer, sheets: Dict) -> None:
        """
        Write main output sheets in the canonical order.

        Pivot2 gets special treatment: two blank rows + title row before
        its actual column headers, to match the expected SAP output format.
        """

        ordering = ["Sheet1", "Pivot1", "Pivot2"]

        for name in ordering:
            frame = sheets.get(name)
            if frame is None:
                continue

            if name == "Pivot2":
                self._write_pivot2_with_header(writer, frame)
            else:
                frame.to_excel(writer, sheet_name=name, index=False)

    def _write_pivot2_with_header(
        self,
        writer,
        frame: pd.DataFrame,
    ) -> None:
        """
        Write Pivot2 with the expected SAP structure:
            Row 1: blank
            Row 2: blank
            Row 3: "Sum of Provision" (title label)
            Row 4: column headers
            Row 5+: data rows

        pandas to_excel does not natively support pre-header rows,
        so we write a spacer dataframe first, then append the real data.
        """

        # Write the three spacer/title rows manually
        spacer = pd.DataFrame(
            [
                [""] * len(frame.columns),     # row 1 — blank
                [""] * len(frame.columns),     # row 2 — blank
                ["Sum of Provision"]            # row 3 — title
                + [""] * (len(frame.columns) - 1),
            ],
            columns=frame.columns,
        )

        # Write spacer (no header row of its own — we'll add the real
        # headers when we write `frame` at startrow=3)
        spacer.to_excel(
            writer,
            sheet_name="Pivot 2",
            index=False,
            header=False,
            startrow=0,
        )

        # Write the actual data with its column headers at row 4 (0-indexed=3)
        frame.to_excel(
            writer,
            sheet_name="Pivot 2",
            index=False,
            startrow=3,
        )

        # ------------------------------------------------------------------
        # Append Grand Total row at the bottom.
        # The ground-truth output ends with:
        #   ['Grand Total', None, None, None, None, None, <sum>]
        # We write this directly onto the worksheet after to_excel finishes.
        # Note: to_excel uses openpyxl under the hood; the sheet already
        # exists in the writer at this point so we can access it via
        # writer.sheets["Pivot 2"].
        # ------------------------------------------------------------------
        ws = writer.sheets["Pivot 2"]
        grand_total = float(frame["Total"].fillna(0).sum())
        last_data_row = ws.max_row + 1  # one row below current last row
        ws.cell(last_data_row, 1).value = "Grand Total"
        ws.cell(last_data_row, 7).value = grand_total

    def _write_warning_sheet(self, writer, dataframe) -> None:
        enabled = self.config.get("workbook", "generate_warning_sheet", True)
        if not enabled:
            return
        if dataframe is None or (hasattr(dataframe, "empty") and dataframe.empty):
            return
        dataframe.to_excel(writer, sheet_name="Warnings", index=False)

    def _write_metadata_sheet(self, writer, dataframe) -> None:
        enabled = self.config.get("workbook", "generate_metadata_sheet", True)
        if not enabled:
            return
        if dataframe is None:
            return
        dataframe.to_excel(writer, sheet_name="Run_Metadata", index=False)

    # ------------------------------------------------------------------
    # Post-write formatting
    # ------------------------------------------------------------------

    def _format_workbook(
        self,
        workbook_path: Path,
        sheets: Dict,
    ) -> None:
        """Apply openpyxl styling to every sheet in the workbook."""

        workbook = load_workbook(workbook_path)

        try:
            for sheet in workbook.worksheets:
                # Determine the header row index (Pivot2 has headers at row 4)
                header_row = 4 if sheet.title == "Pivot 2" else 1

                self._style_header(sheet, header_row)
                self._freeze_header(sheet, header_row)
                self._apply_filter(sheet, header_row)
                self._auto_width(sheet)
                self._number_format(sheet, header_row)
                self._warning_colors(sheet)

            workbook.save(workbook_path)

        finally:
            workbook.close()

    def _style_header(self, worksheet, header_row: int = 1) -> None:
        """Bold, centred, green-fill on the header row."""

        header_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
        )

        for cell in worksheet[header_row]:
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill      = header_fill

    def _freeze_header(self, worksheet, header_row: int = 1) -> None:
        """Freeze panes so the header stays visible while scrolling."""

        if not self.config.get("workbook", "freeze_header_row", True):
            return

        # Freeze below header_row: "A2" for row 1, "A5" for row 4 (Pivot2)
        freeze_cell = f"A{header_row + 1}"
        worksheet.freeze_panes = freeze_cell

    def _apply_filter(self, worksheet, header_row: int = 1) -> None:
        """Add Excel auto-filter dropdowns on the header row."""

        if not self.config.get("workbook", "enable_auto_filter", True):
            return

        # Build a ref that covers only the header row and data below
        max_col_letter = get_column_letter(worksheet.max_column)
        ref = (
            f"A{header_row}:{max_col_letter}{worksheet.max_row}"
        )
        worksheet.auto_filter.ref = ref

    def _auto_width(self, worksheet) -> None:
        """Auto-fit column widths to the widest cell value (max 60 chars)."""

        if not self.config.get("workbook", "auto_column_width", True):
            return

        for col in worksheet.columns:
            max_len = 0
            letter  = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))

            worksheet.column_dimensions[letter].width = min(max_len + 2, 60)

    def _number_format(self, worksheet, header_row: int = 1) -> None:
        """
        Apply "#,##0" number formatting to money columns.

        BUG FIX: original code targeted "business_amount" which is an
        internal column never written to output sheets.
        Fixed to target "Provision" (Sheet1) and "Total" (Pivot1, Pivot2).
        """

        # Build header → column-index map from the actual header row
        headers = {
            cell.value: cell.column
            for cell in worksheet[header_row]
            if cell.value is not None
        }

        for col_name in self.MONEY_COLUMNS:
            col_index = headers.get(col_name)
            if col_index is None:
                continue

            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row_idx, col_index).number_format = "#,##0"

    def _warning_colors(self, worksheet) -> None:
        """Highlight rows in the Warnings sheet with a red fill."""

        if not self.config.get("workbook", "conditional_warning_colors", True):
            return

        if worksheet.title != "Warnings":
            return

        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and "critical" in str(cell.value).lower():
                    cell.fill = red_fill
